import os
from openpyxl import Workbook, load_workbook
from sources.SettingsParser import DATA_PATH

DATA_PATH = DATA_PATH


class XlsRepository:
    def __init__(self, base_file=DATA_PATH, backup_folder=None):
        self.backup_folder = backup_folder
        self.base_file = base_file
        if os.path.exists(self.base_file):
            self.workbook = load_workbook(filename=base_file)
        else:
            self.workbook = Workbook()
            self.drivers_sheet = self.workbook.active
            self.drivers_sheet.title = "Soferi"
            self.routes_sheet = self.workbook.create_sheet("Rute")
            self.workbook.save(self.base_file)

    def write_base_sheet(self, drivers, routes):
        if drivers:
            r = 1
            for driver in drivers:
                _ = self.drivers_sheet.cell(row=r, column=1, value=driver.name)
                r += 1
        if routes:
            r = 1
            for route in routes:
                c = 1
                for slot in route.routes:
                    _ = self.routes_sheet.cell(row=r, column=c, value="{0} {1}".format(slot[1], slot[0]))
                    c += 1
                r += 1
        self.workbook.save(self.base_file)

    def read_drivers(self):
        drivers_name = [cell.value for cell in self.drivers_sheet['A']]
        return drivers_name

    def read_routes(self):
        routes_name=[]
        for row in self.routes_sheet.iter_rows():
            routes_name.append([cell.value for cell in row])
        return routes_name
