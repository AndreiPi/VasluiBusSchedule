import unittest
import os

from sources.XlsRepository import XlsRepository
from sources.Driver import Driver
from sources.Route import Route
from openpyxl import load_workbook


class TestWriteBaseSheets(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.test_drivers = [Driver("Ion"), Driver("Vasile")]
        cls.test_routes = [
            Route(locations=["B1", "R2"], hours=["10:00", "11:00"]),
            Route(locations=["R1", "B2"], hours=["10:00", "11:00"])
        ]
        cls.file_path = "../data/data_test.xlsx"
        cls.repository = XlsRepository(cls.file_path)

    def test_write_file(self):
        self.repository.write_base_sheet(None, None)

        self.test_workbook = load_workbook(filename=self.file_path)
        self.assertIsNotNone(self.test_workbook)

    def test_write_sheets(self):
        sheet_names = ["Soferi", "Rute"]

        self.repository.write_base_sheet(None, None)

        self.test_workbook = load_workbook(filename=self.file_path)
        self.assertEqual(sheet_names, self.test_workbook.sheetnames)

    def test_write_Drivers(self):
        self.repository.write_base_sheet(self.test_drivers, None)

        self.test_workbook = load_workbook(filename=self.file_path)
        drivers_sheet = self.test_workbook["Soferi"]
        drivers_name = [cell.value for cell in drivers_sheet['A']]
        self.assertEqual(["Ion", "Vasile"], drivers_name)

    def test_write_Routes(self):
        self.repository.write_base_sheet(None, self.test_routes)

        self.test_workbook = load_workbook(filename=self.file_path)
        routes_sheet = self.test_workbook["Rute"]
        routes_name = [cell.value for cell in routes_sheet['1']]
        self.assertEqual(["10:00 B1", "11:00 R2"], routes_name)

    def test_read_drivers(self):
        self.repository.write_base_sheet(self.test_drivers, None)

        drivers = self.repository.read_drivers()

        self.assertEqual([x.name for x in self.test_drivers], drivers)

    def test_read_routes(self):
        self.repository.write_base_sheet(None, self.test_routes)

        routes = self.repository.read_routes()

        self.assertEqual(["10:00 B1", "11:00 R2"], routes[0])
        self.assertEqual(["10:00 R1", "11:00 B2"], routes[1])

    @classmethod
    def tearDown(cls):
        os.remove(cls.file_path)


if __name__ == '__main__':
    unittest.main()
